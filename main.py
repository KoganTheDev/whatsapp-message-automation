"""
WhatsApp Message Automation Script
---------------------------------
This script reads contact information and WhatsApp URLs from an Excel file, then sends two text messages (with Hebrew RTL support) and an image to each contact using WhatsApp Desktop automation.

Key Features:
- Reads messages and image path from text/image files.
- Skips already-processed (highlighted) or faulty URL rows.
- Supports Hebrew/Unicode messages using clipboard paste.
- Sends the image in the messages folder using the clipboard.
- Highlights successful rows in Excel for tracking.

Dependencies:
- openpyxl, pyautogui, pyperclip, Pillow (PIL), pywin32
"""
import os
import time
import pyautogui
import pyperclip
import random
import threading
from openpyxl import load_workbook
import win32clipboard
import win32con
from PIL import Image
import io

def load_messages():
    """
    Load the two text messages (with RTL mark for Hebrew) and image path for sending.
    Returns:
        tuple: (first_message_template (str), second_message (str), image_path (str))
    """
    RLM = '\u200F'  # Right-to-Left Mark for Hebrew
    with open('messeges/first_message.txt', 'r', encoding='utf-8') as f:
        first_message_template = RLM + f.read()
    with open('messeges/second_message.txt', 'r', encoding='utf-8') as f:
        second_message = RLM + f.read()
    image_path = os.path.abspath('messeges/image.jpg')
    return first_message_template, second_message, image_path

def send_image_to_clipboard(image_path):
    """
    Copy an image file to the clipboard as an image for pasting in WhatsApp.
    Args:
        image_path (str): Absolute path to the image file.
    """
    image = Image.open(image_path)
    output = io.BytesIO()
    image.convert('RGB').save(output, 'BMP')
    data = output.getvalue()[14:]
    output.close()
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32con.CF_DIB, data)
    win32clipboard.CloseClipboard()


def close_chrome_tab():
    """
    Close the most recently opened Chrome tab (used to close WhatsApp web tab).
    """
    pyautogui.hotkey('alt', 'tab')
    time.sleep(1)
    pyautogui.hotkey('ctrl', 'w')
    pyautogui.hotkey('alt', 'tab')


def detect_phone_not_found(images_dir='images', max_retries=3, delay=1.0):
    """
    Detect if a 'phone number not found' error is present using image recognition.
    Checks for both not_found.png and not_found2.png, with retries and confidence.
    Returns True if found, else False.
    """
    not_found_images = [
        os.path.join(images_dir, 'not_found.png'),
        os.path.join(images_dir, 'not_found2.png')
    ]
    for _ in range(max_retries):
        for img_path in not_found_images:
            if os.path.exists(img_path):
                location = pyautogui.locateOnScreen(img_path, confidence=0.5)
                if location:
                    pyautogui.press('enter')
                    return True
        time.sleep(delay)
    return False


def detect_and_handle_404():
    """
    Detect if a 404 page is open using image recognition. If found, close the tab and return True.
    Returns:
        bool: True if 404 was detected and handled, False otherwise.
    """
    page_404_img = 'images/page_404_whatsapp.png'
    if os.path.exists(page_404_img):
        location = pyautogui.locateOnScreen(page_404_img, confidence=0.8)
        if location:
            pyautogui.hotkey('ctrl', 'w')
            return True
    return False


def set_excel_comment(row, header_map, comment):
    """
    Set a comment in the Comments column for the given row.
    Args:
        row: The row object from openpyxl.
        header_map: Dict mapping header names to indices.
        comment: The comment string to write.
    """
    comments_col = header_map.get('Comments')
    if comments_col is not None:
        try:
            row[comments_col].value = comment
        except Exception as e:
            print(f"Error writing comment to Excel: {e}")


def send_whatsapp_messages(url, first_name, messages):
    """
    Open WhatsApp Desktop with the given URL and send the messages (first message personalized, then second message + image).
    Args:
        url (str): WhatsApp chat URL.
        first_name (str): The first name to personalize the first message.
        messages (tuple): (first_message_template, second_message, image_path)
    Returns:
        str|bool: '404', 'phone_not_found', True, or False
    """
    first_message_template, second_message, image_path = messages
    RLM = '\u200F'
    if first_name is not None:
        personalized_first_message = RLM + f"היי {first_name}!\n" + first_message_template.lstrip(RLM)
    else:
        personalized_first_message = RLM + "היי!\n" + first_message_template.lstrip(RLM)
    combined_message = personalized_first_message + "\n" + second_message
    result = {'status': None}
    def automation():
        try:
            os.startfile(url)
            event = threading.Event()
            event.wait(5)
            if detect_and_handle_404():
                print(f"404 page detected for URL: {url}. Skipping.")
                result['status'] = "404"
                return
            if detect_phone_not_found():
                print(f"Phone number not found for URL: {url}. Skipping.")
                result['status'] = "phone_not_found"
                return
            send_image_to_clipboard(image_path)
            pyautogui.hotkey('ctrl', 'v')  # Paste the image
            event.wait(3)
            # Paste the combined message (first + second)
            pyperclip.copy(combined_message)
            pyautogui.hotkey('ctrl', 'v')
            event.wait(2)
            pyautogui.press('enter')  # Send both messages and image together
            event.wait(0.7)
            close_chrome_tab()  # Close the tab only after sending
            result['status'] = True
        except pyautogui.FailSafeException as fs_err:
            print(f"PyAutoGUI FailSafeException: {fs_err}")
            result['status'] = False
        except Exception as e:
            print(f"Failed to send messages to {url}: {e}")
            result['status'] = False
    t = threading.Thread(target=automation)
    t.start()
    t.join()
    return result['status']


def read_run_state(state_file='run_state.txt'):
    """
    Reads the run state from a text file. Returns (excel_filename, start_row, num_rows).
    If the file does not exist, creates it with defaults and returns them.
    """
    default_excel = 'excel.xlsx'
    default_start = 2  # Excel row numbers (header is row 1)
    default_count = 50
    excel_file = default_excel
    start_row = default_start
    rows_to_process = default_count
    if not os.path.exists(state_file):
        return default_excel, default_start, default_count
    with open(state_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line.startswith('excel_file:'):
                excel_file = line.split(':', 1)[1].strip()
            elif line.startswith('start_row:'):
                start_row = int(line.split(':', 1)[1].strip())
            elif line.startswith('rows_to_process:'):
                rows_to_process = int(line.split(':', 1)[1].strip())
    return excel_file, start_row, rows_to_process


def update_run_state(state_file, excel_filename, next_row, num_rows=None):
    """
    Updates the run state file with the new starting row (and optionally rows_to_process).
    """
    old_rows = 50
    if os.path.exists(state_file):
        with open(state_file, 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('rows_to_process:'):
                    try:
                        old_rows = int(line.split(':', 1)[1].strip())
                    except Exception:
                        pass
    if num_rows is None:
        num_rows = old_rows
    with open(state_file, 'w', encoding='utf-8') as f:
        f.write("# WhatsApp Automation Run State\n")
        f.write(f"excel_file: {excel_filename}\n")
        f.write(f"start_row: {next_row}\n")
        f.write(f"rows_to_process: {num_rows}\n")
        f.write("# Edit these values to control which file, row, and how many rows to process.\n")
        f.write("# The script will update start_row as it progresses.\n")


def main():
    """
    Main function to process the Excel file and send WhatsApp messages.
    - Loads messages and image path.
    - Iterates over each contact in the Excel file.
    - Skips already-processed or faulty URL rows.
    - Sends messages and image via WhatsApp Desktop automation.
    - Marks success in column J.
    - Tracks sent messages in a text file to avoid duplicates.
    - Handles 404 pages and logs comments.
    """
    print("Starting WhatsApp automation script...")
    state_file = 'run_state.txt'
    excel_path, start_row, num_rows = read_run_state(state_file)
    print(f"Excel file: {excel_path}\nStart at row: {start_row}\nRows to process: {num_rows}")
    first_message_template, second_message, image_path = load_messages()
    wb = load_workbook(excel_path)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_map = {cell.value: idx for idx, cell in enumerate(header_row)}
    sent_people_file = 'messeges_sent_to.txt'
    sent_people = set()
    if os.path.exists(sent_people_file):
        with open(sent_people_file, 'r', encoding='utf-8') as f:
            for line in f:
                sent_people.add(line.strip())
    try:
        processed = 0
        for idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            if processed >= num_rows:
                break
            first_name = row[header_map['First Name']].value
            last_name = row[header_map['Last Name']].value
            # Assign url and sent_col before duplicate check
            url = row[header_map['URLs']].value
            sent_col = row[header_map['Message Sent?']]
            # Allow (None, None) to be resent by skipping duplicate check for that case
            is_none_name = (first_name is None and last_name is None)
            person_id = f"{first_name} {last_name}".strip()
            # Duplicate check
            if not is_none_name and person_id in sent_people:
                sent_col.value = "לא"
                set_excel_comment(row, header_map, "Duplicate contact")
                wb.save(excel_path)
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
                continue
            # Faulty URL check
            if url and url.strip() == 'https://wa.me/972':
                sent_col.value = "לא"
                set_excel_comment(row, header_map, "URL is not a real number (https://wa.me/972)")
                wb.save(excel_path)
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
                continue
            # Invalid URL check
            if not (url and url.startswith('https://')):
                sent_col.value = "לא"
                set_excel_comment(row, header_map, "Phone number/URL is not correct")
                wb.save(excel_path)
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
                continue
            # Send WhatsApp messages and handle result
            whatsapp_result = send_whatsapp_messages(
                url, first_name, (first_message_template, second_message, image_path)
            )
            if whatsapp_result == "404":
                sent_col.value = "לא"
                set_excel_comment(row, header_map, "Page not found.")
                wb.save(excel_path)
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
                continue
            elif whatsapp_result == "phone_not_found":
                sent_col.value = "לא"
                set_excel_comment(row, header_map, "Phone number not found.")
                wb.save(excel_path)
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
                continue
            elif whatsapp_result is True:
                sent_col.value = "כן"
                # Only write to sent_people_file if not (None, None)
                if not is_none_name:
                    with open(sent_people_file, 'a', encoding='utf-8') as f:
                        f.write(person_id + '\n')
                    sent_people.add(person_id)
                if (idx - start_row) % 10 == 0:
                    print(f"  Success: {first_name} {last_name} | {url}")
                wb.save(excel_path)
                time.sleep(random.uniform(2, 5))
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
            else:
                sent_col.value = "לא"
                wb.save(excel_path)
                processed += 1
                update_run_state(state_file, excel_path, idx + 1)
    except KeyboardInterrupt:
        print("\nProcess interrupted by user (Ctrl+C). Saving and closing workbook...")
    finally:
        wb.save(excel_path)
        wb.close()
        print("Workbook saved and closed. Process completed.")


if __name__ == "__main__":
    main()